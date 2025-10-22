"""Core implementation for the read-only Excel MCP tools."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


@dataclass
class CellInfo:
    """Container describing a matched Excel cell."""

    row: int
    column: int
    value: Any
    sheet_name: str
    address: str = ""

    def __post_init__(self) -> None:
        if not self.address:
            from openpyxl.utils import get_column_letter

            column_letter = get_column_letter(self.column)
            self.address = f"{column_letter}{self.row}"


@dataclass
class ExcelFileInfo:
    """Summary information for an Excel workbook."""

    file_path: str
    file_size: int
    sheet_count: int
    sheet_names: List[str]


class ExcelReadTools:
    """Collection of Excel read helpers shared by the MCP server."""

    def __init__(self) -> None:
        self.supported_formats = [".xlsx", ".xls", ".xlsm"]

    def _validate_file(self, file_path: Union[str, Path]) -> Path:
        """Verify that the supplied file exists and can be processed."""
        path = Path(file_path)

        if not path.is_absolute():
            raise ValueError("Please provide an absolute path to the Excel file.")

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        if path.suffix.lower() not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {path.suffix}")

        return path

    @staticmethod
    def _convert_datetime_to_string(value: Any) -> str:
        """Normalise datetime-like values to a readable string."""
        try:
            if value is None or pd.isna(value):
                return ""

            if isinstance(value, (datetime, pd.Timestamp)):
                return value.strftime("%Y-%m-%d %H:%M:%S")

            if hasattr(value, "date") and callable(getattr(value, "date")):
                return value.strftime("%Y-%m-%d %H:%M:%S")

            return str(value)
        except Exception as exc:  # pragma: no cover - defensive logging
            logger.warning("Failed to normalise datetime value: %s", exc)
            return str(value) if value is not None else ""

    def _convert_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Return a DataFrame with datetime-like values stringified."""
        if df.empty:
            return df.copy()

        # applymap is safe across pandas versions
        return df.applymap(self._convert_datetime_to_string)

    def excel_read_info(self, file_path: str) -> Dict[str, Any]:
        """Return workbook level metadata."""
        try:
            path = self._validate_file(file_path)
            workbook = openpyxl.load_workbook(path, read_only=True)

            file_info = ExcelFileInfo(
                file_path=str(path),
                file_size=path.stat().st_size,
                sheet_count=len(workbook.sheetnames),
                sheet_names=list(workbook.sheetnames),
            )

            workbook.close()

            return {
                "success": True,
                "data": {
                    "file_path": file_info.file_path,
                    "file_size": file_info.file_size,
                    "sheet_count": file_info.sheet_count,
                    "sheet_names": file_info.sheet_names,
                },
            }
        except Exception as exc:
            logger.error("excel_read_info failed: %s", exc)
            return {"success": False, "error": str(exc)}

    def excel_read_range(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        range_spec: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Read a sheet (optionally a range) and return rows as dicts."""
        try:
            path = self._validate_file(file_path)

            df = pd.read_excel(path, sheet_name=sheet_name)

            if range_spec:
                logger.warning("range_spec is not currently implemented; returning full sheet.")

            converted = self._convert_dataframe(df)

            return {
                "success": True,
                "data": {
                    "sheet_name": sheet_name or getattr(df, "name", "Sheet1"),
                    "shape": converted.shape,
                    "columns": converted.columns.tolist(),
                    "data": converted.to_dict("records"),
                },
            }
        except Exception as exc:
            logger.error("excel_read_range failed: %s", exc)
            return {"success": False, "error": str(exc)}

    def excel_read_all_sheets(
        self,
        file_path: str,
        max_rows_per_sheet: int = 1000,
        include_empty_sheets: bool = False,
    ) -> Dict[str, Any]:
        """Read every sheet in the workbook, optionally truncating rows."""
        try:
            path = self._validate_file(file_path)
            all_sheets = pd.read_excel(path, sheet_name=None)

            sheets_data: Dict[str, Any] = {}
            sheet_summary: List[Dict[str, Any]] = []

            for sheet_name, df in all_sheets.items():
                try:
                    if df.empty and not include_empty_sheets:
                        sheet_summary.append(
                            {
                                "sheet_name": sheet_name,
                                "status": "skipped (empty)",
                                "rows": 0,
                                "columns": 0,
                            }
                        )
                        continue

                    original_rows = len(df)
                    if original_rows > max_rows_per_sheet:
                        df = df.head(max_rows_per_sheet)
                        logger.warning(
                            "Sheet %s truncated to %s rows (original %s)",
                            sheet_name,
                            max_rows_per_sheet,
                            original_rows,
                        )

                    converted = self._convert_dataframe(df)

                    sheets_data[sheet_name] = {
                        "shape": converted.shape,
                        "columns": converted.columns.tolist(),
                        "data": converted.to_dict("records"),
                        "truncated": original_rows > max_rows_per_sheet,
                        "original_rows": original_rows,
                    }

                    sheet_summary.append(
                        {
                            "sheet_name": sheet_name,
                            "status": "loaded",
                            "rows": len(converted),
                            "columns": len(converted.columns),
                            "original_rows": original_rows,
                        }
                    )
                except Exception as sheet_exc:
                    logger.error("Failed to read sheet %s: %s", sheet_name, sheet_exc)
                    sheet_summary.append(
                        {
                            "sheet_name": sheet_name,
                            "status": f"error: {sheet_exc}",
                            "rows": 0,
                            "columns": 0,
                        }
                    )

            return {
                "success": True,
                "data": {
                    "file_path": str(path),
                    "total_sheets": len(all_sheets),
                    "sheet_summary": sheet_summary,
                    "sheets_data": sheets_data,
                    "settings": {
                        "max_rows_per_sheet": max_rows_per_sheet,
                        "include_empty_sheets": include_empty_sheets,
                    },
                },
            }
        except Exception as exc:
            logger.error("excel_read_all_sheets failed: %s", exc)
            return {"success": False, "error": str(exc)}

    def excel_quick_overview(
        self,
        file_path: str,
        sample_rows: int = 5,
    ) -> Dict[str, Any]:
        """Return workbook metadata plus a sample from every sheet."""
        try:
            path = self._validate_file(file_path)
            info = self.excel_read_info(file_path)
            if not info.get("success"):
                return info

            all_sheets = pd.read_excel(path, sheet_name=None)
            overview: List[Dict[str, Any]] = []

            for sheet_name, df in all_sheets.items():
                try:
                    sample_df = df.head(sample_rows) if not df.empty else df
                    converted_sample = self._convert_dataframe(sample_df)

                    overview.append(
                        {
                            "sheet_name": sheet_name,
                            "total_rows": len(df),
                            "total_columns": len(df.columns),
                            "columns": df.columns.tolist(),
                            "sample_data": converted_sample.to_dict("records"),
                            "is_empty": df.empty,
                            "has_more_data": len(df) > sample_rows,
                        }
                    )
                except Exception as sheet_exc:
                    logger.error("Failed to summarise sheet %s: %s", sheet_name, sheet_exc)
                    overview.append(
                        {
                            "sheet_name": sheet_name,
                            "error": str(sheet_exc),
                        }
                    )

            return {
                "success": True,
                "data": {
                    **info["data"],
                    "sheets_overview": overview,
                    "sample_settings": {"sample_rows": sample_rows},
                },
            }
        except Exception as exc:
            logger.error("excel_quick_overview failed: %s", exc)
            return {"success": False, "error": str(exc)}

    def excel_search(
        self,
        file_path: str,
        search_term: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Search for a term across the workbook (or a single sheet)."""
        try:
            path = self._validate_file(file_path)
            results: List[Dict[str, Any]] = []

            if sheet_name:
                df = pd.read_excel(path, sheet_name=sheet_name)
                results.extend(self._search_in_dataframe(df, search_term, sheet_name))
            else:
                excel_file = pd.ExcelFile(path)
                for sheet in excel_file.sheet_names:
                    df = pd.read_excel(path, sheet_name=sheet)
                    results.extend(self._search_in_dataframe(df, search_term, sheet))

            return {
                "success": True,
                "data": {
                    "search_term": search_term,
                    "total_matches": len(results),
                    "matches": results,
                },
            }
        except Exception as exc:
            logger.error("excel_search failed: %s", exc)
            return {"success": False, "error": str(exc)}

    def _search_in_dataframe(
        self,
        df: pd.DataFrame,
        search_term: str,
        sheet_name: str,
    ) -> List[Dict[str, Any]]:
        """Perform a case-insensitive search within a DataFrame."""
        matches: List[Dict[str, Any]] = []
        lowered_term = search_term.lower()

        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row):
                if pd.notna(value) and lowered_term in str(value).lower():
                    converted_value = self._convert_datetime_to_string(value)
                    cell_info = CellInfo(
                        row=row_idx + 2,
                        column=col_idx + 1,
                        value=converted_value,
                        sheet_name=sheet_name,
                    )
                    matches.append(
                        {
                            "sheet_name": cell_info.sheet_name,
                            "address": cell_info.address,
                            "row": cell_info.row,
                            "column": cell_info.column,
                            "value": str(cell_info.value),
                            "column_name": df.columns[col_idx],
                        }
                    )

        return matches
